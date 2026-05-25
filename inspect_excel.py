import openpyxl
import os

# Try multiple locations
possible_paths = [
    r'C:\Users\vince\biocella-deployed2\examples\Copy of USC Microbial Culture Collection (USCMCC).xlsx',
    r'C:\Users\vince\Downloads\Copy of USC Microbial Culture Collection (USCMCC).xlsx',
    os.path.join(os.environ.get('TEMP', ''), 'Copy of USC Microbial Culture Collection (USCMCC).xlsx'),
    os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads', 'Copy of USC Microbial Culture Collection (USCMCC).xlsx')
]

file_path = None
for path in possible_paths:
    if os.path.exists(path):
        file_path = path
        break

if not file_path:
    print(f"File not found. Checked: {possible_paths}")
    exit(1)
wb = openpyxl.load_workbook(file_path, data_only=True)

print('Sheet names:', wb.sheetnames)
print()

for i, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    print(f'=== Sheet {i}: "{sheet_name}" ===')
    print(f'Rows: {len(rows)}, Columns: {len(rows[0]) if rows else 0}')
    print('First 3 rows:')
    for idx, row in enumerate(rows[:3]):
        print(f'  Row {idx}: {row}')
    print()
