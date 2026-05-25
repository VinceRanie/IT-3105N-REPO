import openpyxl
import os

# Try multiple locations
possible_paths = [
    r'C:\Users\vince\biocella-deployed2\examples\Copy of USC Microbial Culture Collection (USCMCC).xlsx',
    r'C:\Users\vince\Downloads\Copy of USC Microbial Culture Collection (USCMCC).xlsx',
]

file_path = None
for path in possible_paths:
    if os.path.exists(path):
        file_path = path
        break

if not file_path:
    print(f"File not found.")
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)

print('Sheet names:', wb.sheetnames)
print()

for i, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    # Count non-empty cells across all rows
    non_empty_count = 0
    non_empty_rows = 0
    for row in rows:
        row_has_data = any(v is not None and str(v).strip() for v in row)
        if row_has_data:
            non_empty_rows += 1
            non_empty_count += sum(1 for v in row if v is not None and str(v).strip())
    
    print(f'=== Sheet {i}: "{sheet_name}" ===')
    print(f'Total rows: {len(rows)}, Non-empty rows: {non_empty_rows}, Total non-empty cells: {non_empty_count}')
    
    # Show first 5 rows with detailed data
    print('First 5 rows (detailed):')
    for idx, row in enumerate(rows[:5]):
        non_empty_in_row = sum(1 for v in row if v is not None and str(v).strip())
        print(f'  Row {idx} ({non_empty_in_row} non-empty): {row[:10]}...')  # Show first 10 columns
    print()
